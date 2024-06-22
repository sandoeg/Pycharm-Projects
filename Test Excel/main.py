from openpyxl import load_workbook

IOS_SFMEA = False  # False would be Android
# SRA
SRA_SEQ_OF_EVENTS_START_ROW = 4
SRA_SEQ_OF_EVENTS_COL = 'C'
# iOS
IOS_START_ROW = 3
IOS_END_ROW = 447
IOS_SEQ_OF_EVENTS_COL = 'B'
IOS_SFMEA_WS_NAME = 'iOS SFMEA'
# Android
ANDROID_START_ROW = 3
ANDROID_END_ROW = 268
ANDROID_SEQ_OF_EVENTS_COL = 'B'
ANDROID_SFMEA_WS_NAME = 'Android SFMEA'
# sFMEA file name
SFMEA_FILE = 'Basal MMA SFMEA.xlsx'


def check_for_matches():
    # sra_rows_to_ignore = [8,16,28,29,30,31,32,33,42,55,61,63,65,66,67,68,69,73,77,80,83,86,87,99,101,103]
    sra_rows_to_ignore = []

    # Initialize iOS and Android data
    if IOS_SFMEA:
        start_row = IOS_START_ROW
        end_row = IOS_END_ROW
        sfmea_seq_of_events_column = IOS_SEQ_OF_EVENTS_COL
        ws_name = IOS_SFMEA_WS_NAME
        sfmea_name = 'iOS'
    else:
        start_row = ANDROID_START_ROW
        end_row = ANDROID_END_ROW
        sfmea_seq_of_events_column = ANDROID_SEQ_OF_EVENTS_COL
        ws_name = ANDROID_SFMEA_WS_NAME
        sfmea_name = 'Android'

    # Open data collection file
    fs = open("SRA to sFMEA Report.txt", "w+")

    # open workbook
    wb = load_workbook(SFMEA_FILE)
    # print(wb.sheetnames)

    # Show all SRA sequence of events
    # SRA Worksheet
    ws_sra = wb["SRA"]
    print("List of SRA Sequence of Events by Column:Row")
    fs.write("List of SRA Sequence of Events by Column:Row\r")

    sra_rows = 0
    for row in range(SRA_SEQ_OF_EVENTS_START_ROW, ws_sra.max_row + 1):
        if row not in sra_rows_to_ignore:
            cell = SRA_SEQ_OF_EVENTS_COL + str(row)
            if ws_sra[cell].value is not None:
                print('\t' + cell + ": " + ws_sra[cell].value)
                fs.write('\t' + cell + ": " + ws_sra[cell].value + "\r")
                sra_rows += 1

    # how many rows in sra
    print("\r")
    print("Number of Active Rows in SRA: " + str(sra_rows))
    print("\r")
    fs.write("\rNumber of Active Rows in SRA: " + str(sra_rows) + "\r")

    print("\r-------------------------------------")
    fs.write("-------------------------------------\r")
    print("List of " + sfmea_name + " Sequence of Events by Column:Row")
    fs.write("List of " + sfmea_name + " Sequence of Events by Column:Row" + "\r")
    # Show all iOS SFMEA sequence of events in sFMEA
    match_cnt = 0
    no_match_cnt = 0
    ws_sfmea = wb[ws_name]
    print("Number of Rows in" + sfmea_name + " sFMEA: " + str(ws_sfmea.max_row))
    fs.write("\rNumber of Rows in" + sfmea_name + " sFMEA: " + str(ws_sfmea.max_row) + "\r")

    sfmea_rows = 0
    for row in range(start_row, ws_sfmea.max_row + 1):
        if row not in sra_rows_to_ignore:
            cell = sfmea_seq_of_events_column + str(row)
            if ws_sfmea[cell].value is not None:
                print("\t" + cell + ": " + ws_sfmea[cell].value)
                fs.write("\t" + cell + ": " + ws_sfmea[cell].value + "\r")
                sfmea_rows += 1

    # Number of active rows in sFMEA
    print("\r")
    print("Number of Active Rows in " + sfmea_name + " sFMEA: " + str(sfmea_rows))
    print("\r")
    fs.write("\rNumber of Active Rows in " + sfmea_name + " sFMEA: " + str(sfmea_rows) + "\r")

    # Find SRA sequence of events matches in the sFMEA

    print("\r\r-----------------Find SRA -> " + sfmea_name + " sFMEA Sequence of Events Matches-----------------\r")
    fs.write("\r-----------------Find SRA -> " + sfmea_name + " sFMEA Sequence of Events Matches-----------------\r")
    for sra_row in range(start_row, ws_sra.max_row + 1):
        if sra_row not in sra_rows_to_ignore:
            sra_cell = SRA_SEQ_OF_EVENTS_COL + str(sra_row)
            if ws_sra[sra_cell].value is not None:
                print("SRA Row: " + str(sra_row) + "\r")
                fs.write("SRA Row: " + str(sra_row) + "\r")
                # search all rows of the sFMEA to see if there is a match
                match_count = False
                for sfmea_row in range(start_row, end_row + 1):
                    sfmea_cell = sfmea_seq_of_events_column + str(sfmea_row)
                    if ws_sfmea[sfmea_cell].value is not None:
                        if ws_sfmea[sfmea_cell].value == ws_sra[sra_cell].value:
                            print("\tMatch found at SRA Row: " + str(sra_row) + "   " + sfmea_name + " sFMEA Row: " + str(
                                sfmea_row))
                            fs.write(
                                "\tMatch found at SRA Row: " + str(sra_row) + "   " + sfmea_name + " Row: " + str(
                                    sfmea_row) + "\r")
                            match_count = True
                            match_cnt += 1
                if not match_count:
                    print("\t*** No " + sfmea_name + " sFMEA match found for SRA Row: " + str(sra_row))
                    fs.write("\t*** No " + sfmea_name + " sFMEA match found for SRA Row: " + str(sra_row) + "\r")
                    no_match_cnt += 1

    # Display and write to file the number of matches
    print("\r")
    print("*** Number of matches of SRA to " + sfmea_name + " sFMEA: " + str(match_cnt))
    fs.write("\r*** Number of matches of SRA to " + sfmea_name + " sFMEA: " + str(match_cnt) + "\r")
    # Display and write to file number of no matches
    print("*** Number of no matches of SRA to " + sfmea_name + " sFMEA: " + str(no_match_cnt))
    fs.write("*** Number of no matches of SRA to " + sfmea_name + " sFMEA: " + str(no_match_cnt) + "\r")

    # close file
    fs.close()


# main()
if __name__ == '__main__':
    check_for_matches()
