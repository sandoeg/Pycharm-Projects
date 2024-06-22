from openpyxl import load_workbook

# Workbooks
IOS_WB: str = 'Basal MMA iOS SFMEA (For Reviewers).xlsx'
ANDROID_WB: str = 'Basal MMA SFMEA.xlsx'
# Worksheets
IOS_WS: str = 'iOS SFMEA'
ANDROID_WS: str = 'Android SFMEA'
# RCM Columns
IOS_RCM_COL: chr = 'E'
ANDROID_RCM_COL: chr = 'E'
# Rows in Worksheet
IOS_START_ROW: int = 2
IOS_END_ROW: int = 308
ANDROID_START_ROW: int = 3
ANDROID_END_ROW: int = 288


def remove_chars(line):
    # local variables
    bad_chars = ['\n\n', ' \n', '\n']

    # Find characters in lines and remove them
    newline = line

    for bad_char in bad_chars:
        newline = line.replace(bad_char, "...")
    """
    if chars_to_remove in line:
        print("there were chars to remove")
    
    if line.count('\n\n'):
        newline = line.replace('\n\n', "  ")
    elif line.count(' \n'):
        newline = line.replace(' \n', "  ")
    elif line.count('\n'):
        newline = line.replace('\n', "  ")
    """

    return newline


def find_rcms():
    # variables
    # Android dictionaries
    android_prelim_dict = {}  # list to hold all android sFMEA RCMs
    android_final_dict = {}  # list to hold non-duplicated android sFMEA RCMs
    # iOS dictionaries
    ios_prelim_dict = {}
    ios_final_dict = {}

    # open workbooks
    wb_ios = load_workbook(IOS_WB)
    # wb_android = load_workbook(ANDROID_WB)

    # open worksheets
    ws_ios = wb_ios[IOS_WS]
    #  = wb_android[ANDROID_WS]

    # Populate the iOS sFMEA RCMs
    for row in range(IOS_START_ROW, IOS_END_ROW + 1):
        cell = IOS_RCM_COL + str(row)

        if ws_ios[cell].value is not None:
            cell_value = remove_chars(ws_ios[cell].value)
            print('\t' + "iOS " + cell + ": " + cell_value)

            if cell_value.count('\n'):
                print("found extra line feed")

            # add this to the list of iOS RCMs
            ios_prelim_dict[cell] = cell_value
    """
    # Remove duplicated RCMs from iOS sFMEA
    # does the final list include the RCM in the prelim list?
    # if not, then add to the final list, otherwise don't add to the final list
    for prelim_list_item in ios_final_dict:
        if prelim_list_item not in ios_final_dict:
            ios_final_list.append(prelim_list_item)

    # Write data to files for review
    fs_prelim = open("iOS Before Removing Dups.txt", "w+")
    fs_final = open('iOS After Removing Dups.txt', "w+")

    fs_final.write("********* After Removing Duplicates *********" + '\r')
    for final_list_item in ios_final_list:
        fs_final.write("Entry: " + final_list_item + '\r')

    fs_prelim.write("\r\r" + "********* Before Removing Duplicates *********" + '\r')
    for prelim_list_item in ios_final_dict:
        fs_prelim.write("Entry: " + prelim_list_item + '\r')

    """


# main()
if __name__ == '__main__':
    find_rcms()
