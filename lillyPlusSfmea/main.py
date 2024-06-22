# imports
from openpyxl import load_workbook

# LillyPlus sFMEA Work Book
LP_SFMEA_WB: str = 'R11 sFMEA - GS Working Copy.xlsx'
# Worksheets
TALTZ_SFMEA_WS: str = 'Taltz SFMEA'
OMVOH_SFMEA_WS: str = 'Omvoh SFMEA'
PHA_WS: str = 'PHA'
# sFMEA Columns
SFMEA_HAZ_SIT_COL: chr = 'K'
SFMEA_P2_COL: chr = "L"
SFMEA_HARMS_COL: chr = "M"
SFMEA_SEVERITY_COL: chr = "O"
SFMEA_HAZARD_COL: chr = "P"
# PHA Columns
PHA_HAZ_SIT_COL: chr = 'B'
PHA_P2_COL: chr = 'C'
PHA_HARMS_COL: chr = 'D'
PHA_SEVERITY_COL: chr = 'E'
PHA_HAZARD_COL: chr = 'A'
# Rows in Worksheet
TALTZ_SFMEA_START_ROW: int = 2
OMVOH_SFMEA_START_ROW: int = 2
PHA_START_ROW: int = 2
# List of Medications
MEDICATION: str = "Taltz"
PHA_MEDICATION_COL: chr = 'F'


def process_sfmea():
    sfmea_wb = load_workbook(LP_SFMEA_WB)
    # open worksheets
    taltz_ws = sfmea_wb[TALTZ_SFMEA_WS]
    pha_ws = sfmea_wb[PHA_WS]

    # Populate the Taltz sFMEA information
    for sfmea_row in range(TALTZ_SFMEA_START_ROW, taltz_ws.max_row + 1):
        cell = SFMEA_HAZ_SIT_COL + str(sfmea_row)

        if taltz_ws[cell].value is not None:
            cell_value = taltz_ws[cell].value
            print("sFMEA Haz Sit " + cell + ": " + cell_value)

            pha_dict = get_pha_data(pha_ws, 'Taltz', cell_value)

            print(pha_dict)
            print('\n')


def get_pha_data(ws, medication, sfmea_haz_sit):
    # Get Hazard, P2, Harm, and Severity from PHA
    # Put the data into the dictionary and return to calling function
    # Find Hazardous Situation match in PHA
    # Iterate through PHA until a match is found for Hazardous Situation

    for row in range(PHA_START_ROW, ws.max_row + 1):
        haz_sit_cell = PHA_HAZ_SIT_COL + str(row)
        medication_cell = PHA_MEDICATION_COL + str(row)

        if ws[haz_sit_cell].value is not None:
            if ws[haz_sit_cell].value == sfmea_haz_sit and ws[medication_cell].value == medication:
                matched_row = row
                pha_haz_sit_cell = PHA_HAZ_SIT_COL + str(matched_row)
                print("PHA Row " + str(matched_row) + " Matching PHA Haz Sit = " + ws[pha_haz_sit_cell].value)
                p2_cell = PHA_P2_COL + str(matched_row)
                harms_cell = PHA_HARMS_COL + str(matched_row)
                severity_cell = PHA_SEVERITY_COL + str(matched_row)
                hazard_cell = PHA_HAZARD_COL + str(matched_row)

                pha_data_dict = {
                    "Haz Sit": ws[pha_haz_sit_cell].value,
                    "P2": ws[p2_cell].value,
                    "Harm": ws[harms_cell].value,
                    "Severity": ws[severity_cell].value,
                    "Hazard": ws[hazard_cell].value

                }
    return pha_data_dict


# main()
if __name__ == '__main__':
    process_sfmea()
